"""
FIFO-filter for fondstransaksjoner
===================================
Filtrerer en transaksjonsliste (Excel) slik at kun transaksjoner som
utgjør dagens beholdning beholdes, basert på FIFO-prinsippet.
Output skrives i ODIN-importformat med tall formatert som tall i Excel.

Transaksjonstyper:
  by/ac/ti  = kjøp (tilgang)
  sl/to     = salg (avgang)
  dv        = utbytte (ignoreres)

Kjøring:
    python fifo_filter.py input.xlsx output.xlsx
"""

import argparse
import io
import sys
from datetime import date
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd


# ── Kandidater for autodeteksjon ───────────────────────────────────────────────
ISIN_KANDIDATER     = ["isin"]
DATO_KANDIDATER     = ["trade date", "tradedate", "trade\ndate", "date", "dato"]
ANDELER_KANDIDATER  = ["quantity", "antall", "andeler", "units", "shares"]
BELOP_KANDIDATER    = ["amount", "beloep", "beløp", "verdi"]
TYPE_KANDIDATER     = ["tran\ncode", "tran code", "trancode", "type", "transtype",
                       "transaction type", "transaksjonstype"]
SECURITY_KANDIDATER      = ["security", "fund class name", "fond", "name", "fondsnavn"]
ORIGINAL_COST_KANDIDATER = ["original cost", "original cost", "originalcost", "cost", "original_cost"]
TRANSFER_KJOP_TYPER      = {"ti", "li"}  # Disse bruker Original Cost hvis tilgjengelig
AVSLUTTENDE_SALG_TYPER   = {"to", "lo"}  # Fjernes fra filtrering hvis de er siste transaksjon per ISIN

KJOP_VERDIER_AUTO = {"by", "kjøp", "kjop", "buy", "purchase", "ac", "ti", "li"}
SALG_VERDIER_AUTO = {"sl", "salg", "sell", "sale", "to", "lo"}
IGNORER_VERDIER   = {"dv", "dividend", "utbytte"}

OUTPUT_KOLONNER = [
    "Customer Name", "Social Security/Organization No", "ODIN Account No",
    "Fund Class Name", "Fund Class ISIN", "Settlement Date", "Shares",
    "Cost Value NAV Date", "Cost Value NOK", "Cost Value Per Share NOK",
    "Cost Value Per Share Interest Part NOK", "Settlement NAV Date",
    "Settlement Currency", "Settlement Amount", "Settlement NAV",
]

# Tallformat for Excel (norsk visning via systeminnstillinger)
FORMAT_ANDELER = '#,##0.0000000'   # 7 desimaler – andeler
FORMAT_BELOP   = '#,##0.00'        # 2 desimaler – beløp


# ── Hjelpefunksjoner ───────────────────────────────────────────────────────────
def til_float(verdi):
    if verdi is None or str(verdi).strip() in ("", "nan", "None"):
        return None
    try:
        return float(Decimal(str(verdi).replace(",", ".")))
    except (InvalidOperation, ValueError):
        return None


def finn_header_rad(filsti: str) -> int:
    df_rå = pd.read_excel(filsti, header=None, dtype=str, nrows=30)
    for i, rad in df_rå.iterrows():
        if rad.astype(str).str.upper().str.strip().eq("ISIN").any():
            print(f"Autodetektert header-rad: {i}")
            return i
    return 0


def les_excel(filsti: str, header_rad: int) -> pd.DataFrame:
    try:
        df = pd.read_excel(filsti, header=header_rad, dtype=str)
    except FileNotFoundError:
        sys.exit(f"Feil: Finner ikke filen '{filsti}'")
    df = df[~df.apply(lambda r: r.astype(str).str.strip().str.match(r"^-+$").all(), axis=1)]
    df = df.reset_index(drop=True)
    print(f"Leste {len(df)} rader fra '{filsti}'")
    print(f"Kolonner: {list(df.columns)}")
    return df


def finn_kolonne_paa_navn(df, kandidater):
    kol_lower = {k.lower().strip(): k for k in df.columns}
    for k in kandidater:
        if k.lower() in kol_lower:
            return kol_lower[k.lower()]
    return None


def finn_type_kolonne(df):
    funnet = finn_kolonne_paa_navn(df, TYPE_KANDIDATER)
    if funnet:
        return funnet
    alle_kjente = KJOP_VERDIER_AUTO | SALG_VERDIER_AUTO
    for kol in df.columns:
        unike = df[kol].dropna().str.strip().str.lower().unique()
        if any(v in alle_kjente for v in unike):
            return kol
    return None


def autodetekter_kolonner(df):
    resultat = {}
    for navn, kandidater, påkrevd in [
        ("isin",     ISIN_KANDIDATER,     True),
        ("dato",     DATO_KANDIDATER,     True),
        ("andeler",  ANDELER_KANDIDATER,  True),
        ("belop",    BELOP_KANDIDATER,    True),
        ("security",      SECURITY_KANDIDATER,      False),
        ("original_cost", ORIGINAL_COST_KANDIDATER, False),
    ]:
        funnet = finn_kolonne_paa_navn(df, kandidater)
        if funnet:
            print(f"  Autodetektert '{navn}': kolonnen '{funnet}'")
            resultat[navn] = funnet
        elif påkrevd:
            sys.exit(f"Feil: Kunne ikke autodetektere '{navn}'.\n"
                     f"  Kolonner: {list(df.columns)}\n"
                     f"  Bruk --{navn} manuelt.")
        else:
            resultat[navn] = None

    type_kol = finn_type_kolonne(df)
    print(f"  Autodetektert 'type': kolonnen '{type_kol}'" if type_kol
          else "  Ingen type-kolonne funnet – bruker fortegn på andeler")
    resultat["type"] = type_kol
    return resultat


def klassifiser_verdier(df, type_kol, kjop_arg=None, salg_arg=None):
    if kjop_arg and salg_arg:
        kjop_set = {v.strip().lower() for v in kjop_arg.split(",")}
        salg_set = {v.strip().lower() for v in salg_arg.split(",")}
        print(f"  Manuell kjøp: {kjop_set}, salg: {salg_set}")
        return kjop_set, salg_set

    unike     = df[type_kol].dropna().str.strip().str.lower().unique()
    kjop_set  = {v for v in unike if v in KJOP_VERDIER_AUTO}
    salg_set  = {v for v in unike if v in SALG_VERDIER_AUTO}
    ignorert  = {v for v in unike if v in IGNORER_VERDIER}
    ukjente   = {v for v in unike if v not in kjop_set | salg_set | ignorert}

    print(f"  Kjøp-typer:  {sorted(kjop_set)}")
    print(f"  Salg-typer:  {sorted(salg_set)}")
    if ignorert: print(f"  Ignoreres:   {sorted(ignorert)}")
    if ukjente:  print(f"  ⚠ Ukjente (ignoreres): {sorted(ukjente)}")

    if not kjop_set:
        sys.exit(f"Feil: Ingen kjente kjøp-verdier i '{type_kol}'.\n"
                 f"  Bruk --kjop (kommaseparert, f.eks. 'by,ac,ti').")
    return kjop_set, salg_set


def forbered_data(df, kol, kjop_set, salg_set, ingen_type):
    df = df.copy()
    df[kol["dato"]] = pd.to_datetime(df[kol["dato"]], dayfirst=True, errors="coerce")
    ugyldige = df[df[kol["dato"]].isna()].index.tolist()
    if ugyldige:
        print(f"Advarsel: {len(ugyldige)} rader med ugyldig dato ignoreres.")
        df = df.dropna(subset=[kol["dato"]])

    def til_decimal(v):
        try:
            return Decimal(str(v).replace(",", ".").strip())
        except Exception:
            return Decimal("0")

    df["_andeler"] = df[kol["andeler"]].apply(til_decimal)
    df["_belop"]   = df[kol["belop"]].apply(til_decimal)

    if ingen_type:
        df["_er_kjop"] = df["_andeler"] > 0
        df["_er_salg"] = df["_andeler"] < 0
    else:
        type_lower     = df[kol["type"]].str.strip().str.lower()
        df["_er_kjop"] = type_lower.isin(kjop_set)
        df["_er_salg"] = type_lower.isin(salg_set)
        df["_andeler"] = df.apply(
            lambda r: abs(r["_andeler"]) if r["_er_kjop"] else
                     -abs(r["_andeler"]) if r["_er_salg"] else r["_andeler"], axis=1)
        df["_belop"]   = df["_belop"].abs()
        df = df[df["_er_kjop"] | df["_er_salg"]].reset_index(drop=True)
        print(f"Etter type-filter: {df['_er_kjop'].sum()} kjøp, {df['_er_salg'].sum()} salg")
    return df


def skill_ut_avsluttende_utforinger(df, kol):
    """
    Finner ISIN-er der to/lo er den siste daterte transaksjonen og
    fjerner alle to/lo-rader for disse ISIN-ene fra filtreringen.
    Returnerer (df_uten_utforinger, liste_av_utforings_info).
    """
    if not kol.get("type") or not kol.get("isin") or not kol.get("dato"):
        return df, []

    df = df.copy()
    df["_dato_tmp"] = pd.to_datetime(df[kol["dato"]], dayfirst=True, errors="coerce")

    utforte = []
    indekser_som_skal_fjernes = []

    for isin, grp in df.groupby(kol["isin"]):
        grp_med_dato = grp.dropna(subset=["_dato_tmp"]).sort_values("_dato_tmp")
        if len(grp_med_dato) == 0:
            continue
        siste_type = str(grp_med_dato.iloc[-1][kol["type"]]).strip().lower()
        if siste_type not in AVSLUTTENDE_SALG_TYPER:
            continue

        to_lo_rader = grp_med_dato[grp_med_dato[kol["type"]].str.strip().str.lower().isin(AVSLUTTENDE_SALG_TYPER)]
        kjop_rader  = grp_med_dato[grp_med_dato[kol["type"]].str.strip().str.lower().isin(KJOP_VERDIER_AUTO)]

        if len(kjop_rader) == 0:
            continue

        siste_kjop_dato  = kjop_rader["_dato_tmp"].max()
        første_tolo_dato = to_lo_rader["_dato_tmp"].min()

        if første_tolo_dato < siste_kjop_dato:
            continue

        def summer(rader, k):
            total = 0
            for v in rader[k]:
                try:
                    total += float(str(v).replace(",", ".").strip())
                except Exception:
                    pass
            return total

        security_navn = ""
        if kol.get("security") and kol["security"] in grp.columns:
            security_navn = str(grp_med_dato.iloc[-1].get(kol["security"], ""))

        utforte.append({
            "isin":         isin,
            "security":     security_navn,
            "type":         siste_type.upper(),
            "antall_rader": len(to_lo_rader),
            "sum_andeler":  summer(to_lo_rader, kol["andeler"]),
            "siste_dato":   grp_med_dato.iloc[-1]["_dato_tmp"].strftime("%d.%m.%Y"),
        })
        indekser_som_skal_fjernes.extend(to_lo_rader.index.tolist())

    df_renset = df.drop(index=indekser_som_skal_fjernes).drop(columns=["_dato_tmp"])
    return df_renset, utforte


def fifo_filter(df, kol):
    df = df.sort_values([kol["isin"], kol["dato"]]).reset_index(drop=True)
    resultat_rader = []

    for isin, gruppe in df.groupby(kol["isin"]):
        kø = []
        for idx, rad in gruppe.iterrows():
            if rad["_er_kjop"]:
                kø.append({"orig_idx": idx, "igjen": rad["_andeler"],
                            "orig_andeler": rad["_andeler"], "orig_belop": rad["_belop"]})
            elif rad["_er_salg"]:
                rest = abs(rad["_andeler"])
                while rest > 0 and kø:
                    første = kø[0]
                    if første["igjen"] <= rest:
                        rest -= første["igjen"]
                        kø.pop(0)
                    else:
                        første["igjen"] -= rest
                        rest = Decimal("0")
                if rest > Decimal("1e-6"):
                    print(f"  Advarsel [{isin}]: salg overskrider kjøp med {rest:.6f}")

        for opp in kø:
            orig = df.loc[opp["orig_idx"]].copy()
            igjen = opp["igjen"]
            if igjen != opp["orig_andeler"] and opp["orig_andeler"] != 0:
                faktor = igjen / opp["orig_andeler"]
                orig[kol["andeler"]] = str(igjen.quantize(Decimal("0.0000001"), rounding=ROUND_HALF_UP))
                orig[kol["belop"]]   = str((opp["orig_belop"] * faktor).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
            # Bruk Original Cost for ti/li hvis tilgjengelig og ikke tom
            if kol.get("original_cost") and kol.get("type"):
                trans_type = str(orig.get(kol["type"], "")).strip().lower()
                if trans_type in TRANSFER_KJOP_TYPER:
                    orig_cost_val = orig.get(kol["original_cost"], "")
                    f = til_float(orig_cost_val)
                    if f is not None and f != 0:
                        orig[kol["belop"]] = str(Decimal(str(f)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
            resultat_rader.append(orig)

    if not resultat_rader:
        print("Ingen gjenværende transaksjoner etter FIFO-filtrering.")
        return pd.DataFrame()
    return pd.DataFrame(resultat_rader).reset_index(drop=True)


def bygg_output_df(df_filtrert, kol):
    """Bygger output-DataFrame med rene streng/Decimal-verdier."""
    today = date.today().strftime("%d.%m.%Y")
    ut = pd.DataFrame(index=df_filtrert.index, columns=OUTPUT_KOLONNER)
    ut["Fund Class Name"]     = df_filtrert[kol["security"]] if kol.get("security") else ""
    ut["Fund Class ISIN"]     = df_filtrert[kol["isin"]]
    ut["Settlement Date"]     = today
    ut["Shares"]              = df_filtrert[kol["andeler"]]
    ut["Cost Value NAV Date"] = pd.to_datetime(df_filtrert[kol["dato"]], errors="coerce").dt.strftime("%d.%m.%Y")
    ut["Cost Value NOK"]      = df_filtrert[kol["belop"]]
    ut["Settlement NAV Date"] = today
    for k in OUTPUT_KOLONNER:
        if k in ut.columns and ut[k].isna().all():
            ut[k] = ""
    return ut[OUTPUT_KOLONNER]


def skriv_excel(df, filsti):
    """Skriver DataFrame til .xlsx med tall som ekte tall og norsk tallformat."""
    wb = openpyxl.Workbook()
    ws = wb.active

    header_font  = Font(bold=True, name="Arial", size=10)
    header_fill  = PatternFill("solid", start_color="D9D9D9")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font    = Font(name="Arial", size=10)

    # Header
    for col_idx, col_navn in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=col_idx, value=col_navn)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    # Data
    for row_idx, rad in enumerate(df.itertuples(index=False), start=2):
        for col_idx, (col_navn, verdi) in enumerate(zip(df.columns, rad), start=1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.font = data_font

            if col_navn == "Shares":
                f = til_float(verdi)
                c.value = f if f is not None else (verdi or None)
                if f is not None:
                    c.number_format = FORMAT_ANDELER
            elif col_navn == "Cost Value NOK":
                f = til_float(verdi)
                c.value = f if f is not None else (verdi or None)
                if f is not None:
                    c.number_format = FORMAT_BELOP
            else:
                c.value = verdi if verdi not in ("", None) else None

    # Kolonnebredde
    for col_idx, col_navn in enumerate(df.columns, start=1):
        max_len = max(
            len(str(col_navn)),
            *[len(str(ws.cell(row=r, column=col_idx).value or ""))
              for r in range(2, ws.max_row + 1)]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    ws.freeze_panes = "A2"
    wb.save(filsti)
    print(f"Lagret {len(df)} rader til '{filsti}'")


def main():
    parser = argparse.ArgumentParser(description="FIFO-filter for fondstransaksjoner")
    parser.add_argument("input")
    parser.add_argument("output")
    parser.add_argument("--isin",       default=None)
    parser.add_argument("--dato",       default=None)
    parser.add_argument("--andeler",    default=None)
    parser.add_argument("--belop",      default=None)
    parser.add_argument("--type",       default=None)
    parser.add_argument("--kjop",       default=None)
    parser.add_argument("--salg",       default=None)
    parser.add_argument("--ingen-type", action="store_true")
    parser.add_argument("--header-rad", type=int, default=None)
    args = parser.parse_args()

    print("\n=== FIFO-filter for fondstransaksjoner ===\n")

    header_rad = args.header_rad if args.header_rad is not None else finn_header_rad(args.input)
    df = les_excel(args.input, header_rad)

    print("\nDetekterer kolonner:")
    auto = autodetekter_kolonner(df) if any(
        v is None for v in [args.isin, args.dato, args.andeler, args.belop]) else {}

    kol = {
        "isin":          args.isin    or auto.get("isin"),
        "dato":          args.dato    or auto.get("dato"),
        "andeler":       args.andeler or auto.get("andeler"),
        "belop":         args.belop   or auto.get("belop"),
        "type":          args.type    or auto.get("type"),
        "security":      auto.get("security"),
        "original_cost": auto.get("original_cost"),
    }

    kjop_set = salg_set = None
    if not args.ingen_type and kol["type"]:
        kjop_set, salg_set = klassifiser_verdier(df, kol["type"], args.kjop, args.salg)
    elif not args.ingen_type and not kol["type"]:
        args.ingen_type = True

    print()
    df, utforte = skill_ut_avsluttende_utforinger(df, kol)
    if utforte:
        print("\n── Avsluttende utføringer fjernet (beholdes i output) ──")
        for u in utforte:
            print(f"  {u['isin']} ({u['security']}): {u['antall_rader']} {u['type']}-rad(er), "
                  f"{u['sum_andeler']:.4f} andeler ført ut {u['siste_dato']}")
    df_klar   = forbered_data(df, kol, kjop_set, salg_set, args.ingen_type)
    df_filtr  = fifo_filter(df_klar, kol)
    df_output = bygg_output_df(df_filtr, kol)
    skriv_excel(df_output, args.output)

    print("\n── Oppsummering per ISIN ──")
    if not df_output.empty:
        for isin, grp in df_output.groupby("Fund Class ISIN"):
            print(f"  {isin}: {len(grp)} transaksjon(er)")
    print(f"\nFerdig. {len(df)} inn → {len(df_output)} ut.\n")


if __name__ == "__main__":
    main()
