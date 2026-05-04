"""
FIFO-filter – Streamlit-app
============================
Start: streamlit run fifo_app.py
Krever: pip install streamlit openpyxl pandas
"""

import io
from datetime import date
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
import streamlit as st

st.set_page_config(page_title="FIFO-filter", page_icon="📊", layout="centered")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;500&display=swap');
html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; }
.stApp { background-color: #0f1117; color: #e8e8e8; }
header[data-testid="stHeader"] { background: transparent; }
.hero { border-left: 3px solid #00d4aa; padding: 1.2rem 1.5rem; margin-bottom: 2rem;
        background: linear-gradient(90deg, rgba(0,212,170,0.06) 0%, transparent 100%); }
.hero h1 { font-family: 'IBM Plex Mono', monospace; font-size: 1.6rem; font-weight: 600;
           color: #00d4aa; margin: 0 0 0.25rem 0; letter-spacing: -0.5px; }
.hero p { font-size: 0.88rem; color: #8a8a9a; margin: 0; font-weight: 300; }
.section-label { font-family: 'IBM Plex Mono', monospace; font-size: 0.7rem; letter-spacing: 2px;
                 text-transform: uppercase; color: #00d4aa; margin-bottom: 0.6rem; margin-top: 1.8rem; }
.stat-row { display: flex; gap: 1rem; margin: 1.2rem 0; }
.stat-card { flex: 1; background: #1a1d27; border: 1px solid #2a2d3a; border-radius: 6px;
             padding: 1rem 1.2rem; text-align: center; }
.stat-card .num { font-family: 'IBM Plex Mono', monospace; font-size: 2rem; font-weight: 600;
                  color: #00d4aa; line-height: 1; }
.stat-card .lbl { font-size: 0.75rem; color: #6a6a7a; margin-top: 0.3rem;
                  text-transform: uppercase; letter-spacing: 1px; }
.stat-card.red .num { color: #ff6b6b; }
.stat-card.neutral .num { color: #e8e8e8; }
.warn-box { background: rgba(255,193,7,0.08); border: 1px solid rgba(255,193,7,0.3);
            border-radius: 6px; padding: 0.8rem 1rem; font-size: 0.85rem;
            color: #ffc107; margin: 0.5rem 0; }
.info-box { background: rgba(0,212,170,0.06); border: 1px solid rgba(0,212,170,0.25);
            border-radius: 6px; padding: 0.8rem 1rem; font-size: 0.82rem;
            color: #8a8a9a; margin: 0.5rem 0; font-family: 'IBM Plex Mono', monospace; }
.stButton > button { background: #00d4aa !important; color: #0f1117 !important;
    font-family: 'IBM Plex Mono', monospace !important; font-weight: 600 !important;
    border: none !important; border-radius: 4px !important; padding: 0.6rem 1.4rem !important;
    font-size: 0.85rem !important; letter-spacing: 0.5px !important; width: 100%; }
.stButton > button:hover { background: #00f0c4 !important; }
.stDownloadButton > button { background: transparent !important; color: #00d4aa !important;
    border: 1px solid #00d4aa !important; font-family: 'IBM Plex Mono', monospace !important;
    font-weight: 600 !important; border-radius: 4px !important; font-size: 0.85rem !important; width: 100%; }
[data-testid="stFileUploader"] { background: #1a1d27; border: 1px dashed #2a2d3a;
                                  border-radius: 8px; padding: 0.5rem; }
[data-testid="stExpander"] { background: #1a1d27; border: 1px solid #2a2d3a; border-radius: 6px; }
hr { border-color: #2a2d3a; }
.stSelectbox > div > div, .stTextInput > div > div > input {
    background: #1a1d27 !important; border-color: #2a2d3a !important; color: #e8e8e8 !important; }
</style>
""", unsafe_allow_html=True)

# ── Konstanter ─────────────────────────────────────────────────────────────────
ISIN_KANDIDATER     = ["isin"]
DATO_KANDIDATER     = ["trade date", "tradedate", "trade\ndate", "date", "dato"]
ANDELER_KANDIDATER  = ["quantity", "antall", "andeler", "units", "shares"]
BELOP_KANDIDATER    = ["amount", "beloep", "beløp", "verdi"]
TYPE_KANDIDATER     = ["tran\ncode", "tran code", "trancode", "type", "transtype",
                       "transaction type", "transaksjonstype"]
SECURITY_KANDIDATER      = ["security", "fund class name", "fond", "name", "fondsnavn"]
ORIGINAL_COST_KANDIDATER = ["original cost", "original cost", "originalcost", "cost", "original_cost"]
TRANSFER_KJOP_TYPER      = {"ti", "li"}  # Disse bruker Original Cost hvis tilgjengelig
AVSLUTTENDE_SALG_TYPER   = {"to", "lo"}  # Disse fjernes fra filtrering hvis de er siste transaksjon per ISIN

KJOP_VERDIER_AUTO = {"by", "kjøp", "kjop", "buy", "purchase", "ac", "ti", "li"}
SALG_VERDIER_AUTO = {"sl", "salg", "sell", "sale", "to", "lo"}
IGNORER_VERDIER   = {"dv", "dividend", "utbytte"}

OUTPUT_KOLONNER = [
    "Customer Name", "Social Security/Organization No", "ODIN Account No",
    "Fund Class Name", "Fund Class ISIN", "Settlement Date", "Shares",
    "Cost Value NAV Date", "Cost Value NOK", "Cost Value Per Share NOK",
    "Cost Value Per Share Interest Part NOK", "Settlement NAV Date",
    "Settlement Currency", "Settlement Amount", "Settlement NAV",
]

FORMAT_ANDELER = '#,##0.0000000'
FORMAT_BELOP   = '#,##0.00'

TYPE_FORKLARING = {
    "by": "kjøp", "sl": "salg", "ac": "acquisition (kjøp)",
    "to": "transfer out (salg)", "ti": "transfer in (kjøp)", "li": "limit in (kjøp)", "lo": "limit out (salg)", "dv": "utbytte (ignoreres)",
}


# ── Hjelpefunksjoner ───────────────────────────────────────────────────────────
def til_float(verdi):
    if verdi is None or str(verdi).strip() in ("", "nan", "None"):
        return None
    try:
        return float(Decimal(str(verdi).replace(",", ".")))
    except (InvalidOperation, ValueError):
        return None

def finn_header_rad(df_rå):
    for i, rad in df_rå.iterrows():
        if rad.astype(str).str.upper().str.strip().eq("ISIN").any():
            return i
    return 0

def finn_kolonne_paa_navn(df, kandidater):
    kol_lower = {k.lower().strip(): k for k in df.columns}
    for k in kandidater:
        if k.lower() in kol_lower:
            return kol_lower[k.lower()]
    return None

def finn_type_kolonne(df):
    funnet = finn_kolonne_paa_navn(df, TYPE_KANDIDATER)
    if funnet:
        return funnet
    alle_kjente = KJOP_VERDIER_AUTO | SALG_VERDIER_AUTO
    for kol in df.columns:
        unike = df[kol].dropna().str.strip().str.lower().unique()
        if any(v in alle_kjente for v in unike):
            return kol
    return None

def autodetekter(df):
    kol = {}
    for navn, kandidater in [
        ("isin", ISIN_KANDIDATER), ("dato", DATO_KANDIDATER),
        ("andeler", ANDELER_KANDIDATER), ("belop", BELOP_KANDIDATER),
        ("security", SECURITY_KANDIDATER),
        ("original_cost", ORIGINAL_COST_KANDIDATER),
    ]:
        kol[navn] = finn_kolonne_paa_navn(df, kandidater)
    kol["type"] = finn_type_kolonne(df)
    return kol

def klassifiser_verdier(df, type_kol, kjop_override=None, salg_override=None):
    if kjop_override and salg_override:
        return ({v.strip().lower() for v in kjop_override},
                {v.strip().lower() for v in salg_override}, set())
    unike     = df[type_kol].dropna().str.strip().str.lower().unique()
    kjop_set  = {v for v in unike if v in KJOP_VERDIER_AUTO}
    salg_set  = {v for v in unike if v in SALG_VERDIER_AUTO}
    ignor_set = {v for v in unike if v in IGNORER_VERDIER}
    ukjente   = {v for v in unike if v not in kjop_set | salg_set | ignor_set}
    return kjop_set, salg_set, ukjente

def skill_ut_avsluttende_utforinger(df, kol):
    """
    Finner ISIN-er der to/lo er den siste daterte transaksjonen og
    fjerner alle to/lo-rader for disse ISIN-ene fra filtreringen.
    Returnerer (df_uten_utforinger, liste_av_utforings_info).
    """
    if not kol.get("type") or not kol.get("isin") or not kol.get("dato"):
        return df, []

    df = df.copy()
    df["_dato_tmp"] = pd.to_datetime(df[kol["dato"]], dayfirst=True, errors="coerce")

    utforte = []
    indekser_som_skal_fjernes = []

    for isin, grp in df.groupby(kol["isin"]):
        grp_med_dato = grp.dropna(subset=["_dato_tmp"]).sort_values("_dato_tmp")
        if len(grp_med_dato) == 0:
            continue
        siste_type = str(grp_med_dato.iloc[-1][kol["type"]]).strip().lower()
        if siste_type not in AVSLUTTENDE_SALG_TYPER:
            continue

        # Sjekk at alle to/lo kommer etter alle kjøp → avsluttende utføring
        to_lo_rader = grp_med_dato[grp_med_dato[kol["type"]].str.strip().str.lower().isin(AVSLUTTENDE_SALG_TYPER)]
        kjop_rader  = grp_med_dato[grp_med_dato[kol["type"]].str.strip().str.lower().isin(KJOP_VERDIER_AUTO)]

        if len(kjop_rader) == 0:
            continue

        siste_kjop_dato  = kjop_rader["_dato_tmp"].max()
        første_tolo_dato = to_lo_rader["_dato_tmp"].min()

        if første_tolo_dato < siste_kjop_dato:
            # to/lo er blandet inn i kjøpsrekken – ikke en ren avsluttende utføring
            continue

        # Beregn total andeler og beløp for to/lo-radene
        def summer_kolonne(rader, kol_navn):
            total = 0
            for v in rader[kol_navn]:
                try:
                    total += float(str(v).replace(",", ".").strip())
                except Exception:
                    pass
            return total

        security_navn = ""
        if kol.get("security") and kol["security"] in grp.columns:
            security_navn = str(grp_med_dato.iloc[-1].get(kol["security"], ""))

        utforte.append({
            "isin":        isin,
            "security":    security_navn,
            "type":        siste_type.upper(),
            "antall_rader": len(to_lo_rader),
            "sum_andeler": summer_kolonne(to_lo_rader, kol["andeler"]),
            "siste_dato":  grp_med_dato.iloc[-1]["_dato_tmp"].strftime("%d.%m.%Y"),
        })
        indekser_som_skal_fjernes.extend(to_lo_rader.index.tolist())

    df_renset = df.drop(index=indekser_som_skal_fjernes).drop(columns=["_dato_tmp"])
    return df_renset, utforte


def les_fil(fil):
    df_rå = pd.read_excel(fil, header=None, dtype=str, nrows=30)
    header_rad = finn_header_rad(df_rå)
    fil.seek(0)
    df = pd.read_excel(fil, header=header_rad, dtype=str)
    df = df[~df.apply(lambda r: r.astype(str).str.strip().str.match(r"^-+$").all(), axis=1)]
    return df.reset_index(drop=True), header_rad

def kjor_fifo(df, kol, kjop_set, salg_set):
    df = df.copy()
    advarsler = []

    df[kol["dato"]] = pd.to_datetime(df[kol["dato"]], dayfirst=True, errors="coerce")
    ugyldige = df[df[kol["dato"]].isna()].index.tolist()
    if ugyldige:
        advarsler.append(f"{len(ugyldige)} rad(er) med ugyldig dato ble ignorert.")
        df = df.dropna(subset=[kol["dato"]])

    def til_decimal(v):
        try:
            return Decimal(str(v).replace(",", ".").strip())
        except Exception:
            return Decimal("0")

    df["_andeler"] = df[kol["andeler"]].apply(til_decimal)
    df["_belop"]   = df[kol["belop"]].apply(til_decimal)

    # Fjern avsluttende to/lo-utføringer før FIFO-filtrering
    df, utforte = skill_ut_avsluttende_utforinger(df, kol)

    ingen_type = kol["type"] is None
    if not ingen_type:
        type_lower     = df[kol["type"]].str.strip().str.lower()
        df["_er_kjop"] = type_lower.isin(kjop_set)
        df["_er_salg"] = type_lower.isin(salg_set)
        df["_andeler"] = df.apply(
            lambda r: abs(r["_andeler"]) if r["_er_kjop"] else
                     -abs(r["_andeler"]) if r["_er_salg"] else r["_andeler"], axis=1)
        df["_belop"]   = df["_belop"].abs()
        df = df[df["_er_kjop"] | df["_er_salg"]].reset_index(drop=True)
    else:
        df["_er_kjop"] = df["_andeler"] > 0
        df["_er_salg"] = df["_andeler"] < 0

    df = df.sort_values([kol["isin"], kol["dato"]]).reset_index(drop=True)
    resultat_rader = []

    for isin, gruppe in df.groupby(kol["isin"]):
        kø = []
        for idx, rad in gruppe.iterrows():
            if rad["_er_kjop"]:
                kø.append({"orig_idx": idx, "igjen": rad["_andeler"],
                            "orig_andeler": rad["_andeler"], "orig_belop": rad["_belop"]})
            elif rad["_er_salg"]:
                rest = abs(rad["_andeler"])
                while rest > 0 and kø:
                    første = kø[0]
                    if første["igjen"] <= rest:
                        rest -= første["igjen"]
                        kø.pop(0)
                    else:
                        første["igjen"] -= rest
                        rest = Decimal("0")
                if rest > Decimal("1e-6"):
                    advarsler.append(f"{isin}: salg overskrider kjøp med {rest:.6f} andeler.")

        for opp in kø:
            orig = df.loc[opp["orig_idx"]].copy()
            igjen = opp["igjen"]
            if igjen != opp["orig_andeler"] and opp["orig_andeler"] != 0:
                faktor = igjen / opp["orig_andeler"]
                orig[kol["andeler"]] = str(igjen.quantize(Decimal("0.0000001"), rounding=ROUND_HALF_UP))
                orig[kol["belop"]]   = str((opp["orig_belop"] * faktor).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
            # Bruk Original Cost for ti/li hvis tilgjengelig og ikke tom
            if kol.get("original_cost") and kol.get("type"):
                trans_type = str(orig.get(kol["type"], "")).strip().lower()
                if trans_type in TRANSFER_KJOP_TYPER:
                    orig_cost_val = orig.get(kol["original_cost"], "")
                    f = til_float(orig_cost_val)
                    if f is not None and f != 0:
                        orig[kol["belop"]] = str(Decimal(str(f)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
            resultat_rader.append(orig)

    if not resultat_rader:
        return pd.DataFrame(columns=OUTPUT_KOLONNER), advarsler, utforte

    df_filtr = pd.DataFrame(resultat_rader).reset_index(drop=True)
    today = date.today().strftime("%d.%m.%Y")
    ut = pd.DataFrame(index=df_filtr.index, columns=OUTPUT_KOLONNER)
    ut["Fund Class Name"]     = df_filtr[kol["security"]] if kol.get("security") else ""
    ut["Fund Class ISIN"]     = df_filtr[kol["isin"]]
    ut["Settlement Date"]     = today
    ut["Shares"]              = df_filtr[kol["andeler"]]
    ut["Cost Value NAV Date"] = pd.to_datetime(df_filtr[kol["dato"]], errors="coerce").dt.strftime("%d.%m.%Y")
    ut["Cost Value NOK"]      = df_filtr[kol["belop"]]
    ut["Settlement NAV Date"] = today
    for k in OUTPUT_KOLONNER:
        if k in ut.columns and ut[k].isna().all():
            ut[k] = ""
    return ut[OUTPUT_KOLONNER], advarsler, utforte


def df_til_excel_bytes(df):
    """Skriver DataFrame til Excel med tall som ekte tall og norsk tallformat."""
    wb = openpyxl.Workbook()
    ws = wb.active

    header_font  = Font(bold=True, name="Arial", size=10)
    header_fill  = PatternFill("solid", start_color="D9D9D9")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font    = Font(name="Arial", size=10)

    for col_idx, col_navn in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=col_idx, value=col_navn)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    for row_idx, rad in enumerate(df.itertuples(index=False), start=2):
        for col_idx, (col_navn, verdi) in enumerate(zip(df.columns, rad), start=1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.font = data_font
            if col_navn == "Shares":
                f = til_float(verdi)
                c.value = f if f is not None else (verdi or None)
                if f is not None:
                    c.number_format = FORMAT_ANDELER
            elif col_navn == "Cost Value NOK":
                f = til_float(verdi)
                c.value = f if f is not None else (verdi or None)
                if f is not None:
                    c.number_format = FORMAT_BELOP
            else:
                c.value = verdi if verdi not in ("", None) else None

    for col_idx, col_navn in enumerate(df.columns, start=1):
        max_len = max(
            len(str(col_navn)),
            *[len(str(ws.cell(row=r, column=col_idx).value or ""))
              for r in range(2, ws.max_row + 1)]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── UI ─────────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="hero">
  <h1>FIFO-filter</h1>
  <p>Filtrer fondstransaksjoner og eksporter i ODIN-importformat</p>
</div>
""", unsafe_allow_html=True)

st.markdown('<div class="section-label">① Last opp transaksjonsfil</div>', unsafe_allow_html=True)
opplastet = st.file_uploader("", type=["xlsx", "xls"], label_visibility="collapsed")

if opplastet:
    try:
        df_original, header_rad = les_fil(opplastet)
    except Exception as e:
        st.error(f"Kunne ikke lese filen: {e}")
        st.stop()

    kol_auto = autodetekter(df_original)
    alle_kolonner = list(df_original.columns)

    st.markdown('<div class="section-label">② Kolonner</div>', unsafe_allow_html=True)
    with st.expander("Autodetekterte kolonner – klikk for å justere", expanded=False):
        c1, c2 = st.columns(2)
        with c1:
            kol_isin    = st.selectbox("ISIN", alle_kolonner,
                index=alle_kolonner.index(kol_auto["isin"]) if kol_auto["isin"] in alle_kolonner else 0)
            kol_dato    = st.selectbox("Dato", alle_kolonner,
                index=alle_kolonner.index(kol_auto["dato"]) if kol_auto["dato"] in alle_kolonner else 0)
            sek_alt     = ["(ingen)"] + alle_kolonner
            kol_security = st.selectbox("Fondsnavn", sek_alt,
                index=(alle_kolonner.index(kol_auto["security"]) + 1) if kol_auto["security"] in alle_kolonner else 0)
        with c2:
            kol_andeler = st.selectbox("Antall andeler", alle_kolonner,
                index=alle_kolonner.index(kol_auto["andeler"]) if kol_auto["andeler"] in alle_kolonner else 0)
            kol_belop   = st.selectbox("Beløp", alle_kolonner,
                index=alle_kolonner.index(kol_auto["belop"]) if kol_auto["belop"] in alle_kolonner else 0)

        st.markdown("**Transaksjonstype**")
        ingen_type = st.checkbox("Ingen type-kolonne (bruk fortegn på andeler)",
                                 value=kol_auto["type"] is None)
        kol_type = None
        if not ingen_type:
            type_alt = ["(ingen)"] + alle_kolonner
            type_def = (alle_kolonner.index(kol_auto["type"]) + 1
                        if kol_auto["type"] and kol_auto["type"] in alle_kolonner else 0)
            kol_type_valg = st.selectbox("Type-kolonne", type_alt, index=type_def)
            kol_type = kol_type_valg if kol_type_valg != "(ingen)" else None

    kol_valgt = {
        "isin":          kol_isin,
        "dato":          kol_dato,
        "andeler":       kol_andeler,
        "belop":         kol_belop,
        "type":          kol_type,
        "security":      kol_security if kol_security != "(ingen)" else None,
        "original_cost": kol_auto.get("original_cost"),
    }

    kjop_set = salg_set = set()
    if kol_type and kol_type in df_original.columns:
        kjop_set, salg_set, ukjente = klassifiser_verdier(df_original, kol_type)
        unike_alle = sorted(df_original[kol_type].dropna().str.strip().unique())

        st.markdown('<div class="section-label">③ Transaksjonstyper</div>', unsafe_allow_html=True)
        type_info = []
        for v in unike_alle:
            v_lower = v.lower()
            if v_lower in kjop_set:
                type_info.append(f"<b style='color:#00d4aa'>{v}</b> → {TYPE_FORKLARING.get(v_lower,'kjøp')}")
            elif v_lower in salg_set:
                type_info.append(f"<b style='color:#ff6b6b'>{v}</b> → {TYPE_FORKLARING.get(v_lower,'salg')}")
            elif v_lower in IGNORER_VERDIER:
                type_info.append(f"<span style='color:#555'>{v}</span> → ignoreres")
            else:
                type_info.append(f"<b style='color:#ffc107'>{v}</b> → ukjent")
        st.markdown('<div class="info-box">' + "&nbsp;&nbsp;|&nbsp;&nbsp;".join(type_info) + "</div>",
                    unsafe_allow_html=True)
        if ukjente:
            st.markdown(f'<div class="warn-box">⚠ Ukjente transaksjonstyper: {", ".join(sorted(ukjente))} – ignoreres.</div>',
                        unsafe_allow_html=True)
        step = "④"
    else:
        step = "③"

    st.markdown(f'<div class="section-label">{step} Forhåndsvisning</div>', unsafe_allow_html=True)
    st.dataframe(df_original.head(10), use_container_width=True, height=220)
    st.caption(f"Viser 10 av {len(df_original)} rader · Header på rad {header_rad}")

    next_step = chr(ord(step[0]) + 1)
    st.markdown(f'<div class="section-label">{next_step} Kjør filter</div>', unsafe_allow_html=True)
    if st.button("▶  Kjør FIFO-filter"):
        with st.spinner("Prosesserer…"):
            try:
                resultat, advarsler, utforte = kjor_fifo(df_original, kol_valgt, kjop_set, salg_set)
                st.session_state["resultat"]  = resultat
                st.session_state["advarsler"] = advarsler
                st.session_state["utforte"]   = utforte
                st.session_state["n_inn"]     = len(df_original)
            except Exception as e:
                st.error(f"Feil under prosessering: {e}")

    if "resultat" in st.session_state:
        resultat  = st.session_state["resultat"]
        advarsler = st.session_state["advarsler"]
        utforte   = st.session_state.get("utforte", [])
        n_inn     = st.session_state["n_inn"]
        n_ut      = len(resultat)

        st.markdown("---")
        st.markdown('<div class="section-label">Resultat</div>', unsafe_allow_html=True)
        st.markdown(f"""
        <div class="stat-row">
          <div class="stat-card neutral"><div class="num">{n_inn}</div><div class="lbl">Inn</div></div>
          <div class="stat-card red"><div class="num">{n_inn - n_ut}</div><div class="lbl">Fjernet</div></div>
          <div class="stat-card"><div class="num">{n_ut}</div><div class="lbl">Beholdt</div></div>
        </div>
        """, unsafe_allow_html=True)

        for a in advarsler:
            st.markdown(f'<div class="warn-box">⚠ {a}</div>', unsafe_allow_html=True)

        st.dataframe(resultat, use_container_width=True, height=350)

        if utforte:
            st.markdown(
                '<div class="warn-box">⚠ <b>Avsluttende utføringer fjernet fra filtrering</b><br>'
                'Følgende ISIN-er hadde to/lo-transaksjoner som avslutter hele beholdningen. '
                'Disse er fjernet fra FIFO-filtreringen, og beholdningen før utføringen beholdes i output.</div>',
                unsafe_allow_html=True
            )
            utforte_df = pd.DataFrame(utforte).rename(columns={
                "isin":         "Fund Class ISIN",
                "security":     "Verdipapir",
                "type":         "Type",
                "antall_rader": "Antall rader fjernet",
                "sum_andeler":  "Sum andeler ført ut",
                "siste_dato":   "Dato siste utføring",
            })
            utforte_df["Sum andeler ført ut"] = utforte_df["Sum andeler ført ut"].round(7)
            st.dataframe(utforte_df, use_container_width=True, hide_index=True)

        with st.expander("Oppsummering per ISIN"):
            opps = (
                resultat.groupby(["Fund Class ISIN", "Fund Class Name"])
                .agg(
                    Transaksjoner=("Fund Class ISIN", "size"),
                    Sum_andeler=("Shares", lambda x: sum(
                        float(str(v).replace(",", ".")) for v in x
                        if str(v).strip() not in ("", "nan", "None")
                    ))
                )
                .reset_index()
                .rename(columns={
                    "Fund Class Name": "Verdipapir",
                    "Sum_andeler": "Sum andeler"
                })
            )
            opps["Sum andeler"] = opps["Sum andeler"].round(7)
            st.dataframe(opps, use_container_width=True, hide_index=True)

        st.markdown('<div class="section-label">Last ned</div>', unsafe_allow_html=True)
        navn = opplastet.name
        if navn.endswith(".xlsx"):
            filnavn_ut = navn[:-5] + "_FIFO.xlsx"
        elif navn.endswith(".xls"):
            filnavn_ut = navn[:-4] + "_FIFO.xlsx"
        else:
            filnavn_ut = navn + "_FIFO.xlsx"
        st.download_button(
            label="⬇  Last ned i ODIN-importformat",
            data=df_til_excel_bytes(resultat),
            file_name=filnavn_ut,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

else:
    st.markdown("""
    <div style="background:#1a1d27;border:1px dashed #2a2d3a;border-radius:8px;
                padding:2.5rem;text-align:center;color:#4a4a5a;font-size:0.9rem;margin-top:1rem;">
        Last opp en Excel-fil ovenfor for å komme i gang
    </div>
    """, unsafe_allow_html=True)
